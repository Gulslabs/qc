"""
Append one row per Naqeeb (from 'Reference Data') into the
'Combined_Report' sheet of KAR_Naqeeb_Amaal_And_Seerah_Tracker_Master_Sheet.xlsx,
for a given Week_End_Date / Class_Date, pulling values from a weekly
Source Excel ('Askaar', 'Halqa', 'Reference Data' sheets).

Usage:
    python append_naqeeb_report.py <source_file.xlsx> <week_end_date> <class_date> \
        [--output OUTPUT_FILE.xlsx]

Dates can be given as DD-Mon-YY, DD-Mon-YYYY, DD-MM-YYYY, DD/MM/YYYY, YYYY-MM-DD.
Example:
    python append_naqeeb_report.py weekly_source.xlsx 09-Aug-26 08-08-2026
"""

import argparse
import difflib
import re
from copy import copy
from datetime import datetime

import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import range_boundaries

BHAI_RE = re.compile(r"\s*bhai\s*$", re.IGNORECASE)
DATE_FORMATS = ("%d-%b-%y", "%d-%b-%Y", "%d/%m/%Y", "%m/%d/%Y", "%Y-%m-%d", "%d-%m-%Y")


# --------------------------------------------------------------------------- #
# Helpers
# --------------------------------------------------------------------------- #
def normalize_name(name):
    """Lowercase, strip trailing 'Bhai', drop punctuation, collapse spaces."""
    if not name:
        return ""
    name = BHAI_RE.sub("", str(name).strip())
    name = re.sub(r"[^a-z0-9 ]", "", name.lower())
    return re.sub(r"\s+", " ", name).strip()


def parse_date(value):
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value.date()
    if hasattr(value, "year") and hasattr(value, "month"):
        return value
    s = str(value).strip()
    for fmt in DATE_FORMATS:
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


def read_sheet_as_dicts(ws):
    headers = [c.value for c in ws[1]]
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if all(v is None for v in row):
            continue
        rows.append(dict(zip(headers, row)))
    return rows


def fuzzy_pick(target_norm, candidates_norm, cutoff=0.72):
    match = difflib.get_close_matches(target_norm, candidates_norm, n=1, cutoff=cutoff)
    return match[0] if match else None


def find_row(naqeeb_name, date_value, pool, name_field, date_field, sheet_label, log):
    """Find the row for naqeeb_name on date_value within pool, exact then fuzzy."""
    target = normalize_name(naqeeb_name)
    date_value = parse_date(date_value)
    subset = [r for r in pool if parse_date(r.get(date_field)) == date_value]
    names_norm = [normalize_name(r.get(name_field)) for r in subset]

    if target in names_norm:
        return subset[names_norm.index(target)]

    best = fuzzy_pick(target, names_norm)
    if best:
        matched_row = subset[names_norm.index(best)]
        log.append(
            f"    [{sheet_label}] fuzzy-matched '{naqeeb_name}' -> "
            f"'{matched_row.get(name_field)}'  (verify this)"
        )
        return matched_row

    log.append(f"    [{sheet_label}] NO MATCH for '{naqeeb_name}' on {date_value}")
    return None


def find_header_row(ws, anchor="Naqeeb Name", search_rows=10):
    for r in range(1, search_rows + 1):
        if ws.cell(row=r, column=1).value == anchor:
            return r
    raise ValueError(f"Could not locate header row (col A == '{anchor}') in first {search_rows} rows")


def build_header_map(ws, header_row):
    headers = {}
    for c in range(1, ws.max_column + 1):
        v = ws.cell(row=header_row, column=c).value
        if v:
            headers[str(v).strip()] = c
    return headers


def find_last_data_row(ws, header_row, name_col):
    last_row = header_row
    for r in range(header_row + 1, ws.max_row + 1):
        if ws.cell(row=r, column=name_col).value not in (None, ""):
            last_row = r
    return last_row


# --------------------------------------------------------------------------- #
# Main
# --------------------------------------------------------------------------- #
def main():
    parser = argparse.ArgumentParser(description="Append weekly Naqeeb rows to Combined_Report")
    parser.add_argument("source_file", help="Path to the weekly Source Excel")
    parser.add_argument("week_end_date", help="Week Ending Date, e.g. 09-Aug-26")
    parser.add_argument("class_date", help="Halqa Class Date, e.g. 08-08-2026")
    parser.add_argument(
        "--output",
        default="KAR_Naqeeb_Amaal_And_Seerah_Tracker_Master_Sheet.xlsx",
        help="Master output file (edited in place)",
    )
    args = parser.parse_args()

    log = []

    # ---- load source ----
    src = openpyxl.load_workbook(args.source_file, data_only=True)
    ref_rows = read_sheet_as_dicts(src["Reference Data"])
    askaar_rows = read_sheet_as_dicts(src["Askaar"])
    halqa_rows = read_sheet_as_dicts(src["Halqa"])

    # ---- load & locate output sheet ----
    wb = openpyxl.load_workbook(args.output)
    ws = wb["Combined_Report"]
    header_row = find_header_row(ws)
    headers = build_header_map(ws, header_row)

    def col(name):
        return get_column_letter(headers[name])

    def set_cell(row, name, value):
        c = headers.get(name)
        if c:
            ws.cell(row=row, column=c, value=value)
        else:
            log.append(f"    [WARN] column '{name}' not found in sheet; skipped")

    last_row = find_last_data_row(ws, header_row, headers["Naqeeb Name"])
    next_row = last_row + 1
    week_end_date_val = parse_date(args.week_end_date)

    # locate the table (if any) up front so we know its column bounds
    table = next(iter(ws.tables.values()), None)
    if table:
        t_min_col, t_min_row, t_max_col, t_max_row = range_boundaries(table.ref)
    else:
        t_min_col, t_max_col = 1, ws.max_column

    def copy_row_style(template_row, target_row, min_col, max_col):
        """Copy cell formatting (number format, font, border, fill) from an
        existing data row onto a newly appended row, column by column."""
        for c in range(min_col, max_col + 1):
            src_cell = ws.cell(row=template_row, column=c)
            dst_cell = ws.cell(row=target_row, column=c)
            dst_cell._style = copy(src_cell._style)

    # ---- append one row per Reference Data.Naqeeb Name ----
    for i, ref in enumerate(ref_rows):
        naqeeb_name = ref.get("Naqeeb Name")
        if not naqeeb_name:
            continue
        row = next_row + i
        log.append(f"Row {row}: {naqeeb_name}")

        copy_row_style(last_row, row, t_min_col, t_max_col)

        askaar = find_row(naqeeb_name, args.week_end_date, askaar_rows,
                           "Generated Naqeeb Name", "Week Ending Date", "Askaar", log)

        # Halqa sheet only contains rows for Naqeebs whose Reference Data role
        # is 'Class-Naqeeb' — skip the lookup entirely for everyone else so a
        # normal absence from Halqa isn't logged as a data problem.
        role = ref.get("Current Role")
        if role == "Class-Naqeeb":
            halqa = find_row(naqeeb_name, args.class_date, halqa_rows,
                              "Naqeeb Name", "Date", "Halqa", log)
        else:
            halqa = None

        set_cell(row, "Naqeeb Name", naqeeb_name)
        set_cell(row, "Role", ref.get("Current Role"))
        
        # Category is now formula-based on Overall Percentage thresholds
        if "Category" in headers and "Overall  Percentage(100%)" in headers:
            set_cell(row, "Category",
                     f'=IFERROR(IF({col("Overall  Percentage(100%)")}{row}>=0.91,"L1",'
                     f'IF({col("Overall  Percentage(100%)")}{row}>=0.8,"L2","L3")),"L3")')
        else:
            set_cell(row, "Category", ref.get("Category"))

        if askaar:
            roza = askaar.get("Roza")
            if isinstance(roza, (int, float)) and roza >= 2:
                roza = 1
            set_cell(row, "Ba-Wazu", askaar.get("Ba-Wuzu"))
            set_cell(row, "Tahajjud", askaar.get("Tahajjud"))
            set_cell(row, "Azkaar", askaar.get("Morning Azkaar"))
            set_cell(row, "Roza", roza)
            set_cell(row, "Sadqa", askaar.get("Sadaqah"))
            set_cell(row, "Listened to Previous Tarbiya Audio",
                     askaar.get("Did you listen to previous Tarbiya Class Recording?"))

        set_cell(row, "Read Halqa Notes(<48 Hours)", "Y")

        if halqa:
            set_cell(row, "No of Students", halqa.get("Total Active"))
            set_cell(row, "Present", halqa.get("TP"))
            set_cell(row, "Absent", halqa.get("A"))
            set_cell(row, "Leave", halqa.get("L"))
            set_cell(row, "Drop in Halqa", halqa.get("Halqa Abs"))

        set_cell(row, "Week Ending Date", week_end_date_val)

        # ---- formula columns ----
        if {"Ba-Wazu", "Sadqa", "Ruju Percentage (33.33%)"} <= headers.keys():
            set_cell(row, "Ruju Percentage (33.33%)",
                     f"=SUM({col('Ba-Wazu')}{row}:{col('Sadqa')}{row})/30*0.33")

        if {"Book Completion", "Listened to Previous Tarbiya Audio", "Role",
            "Read Halqa Notes(<48 Hours)", "Fikr Percentage (33.33%)"} <= headers.keys():
            set_cell(row, "Fikr Percentage (33.33%)",
                     f'=(({col("Book Completion")}{row}*1/3) + '
                     f'IF({col("Listened to Previous Tarbiya Audio")}{row}="Y",1/3,0) + '
                     f'IF(OR({col("Role")}{row}<>"Class-Naqeeb",'
                     f'{col("Read Halqa Notes(<48 Hours)")}{row}="Y"),1/3,0))*0.33')

        if "Attendance Percentage (33.33%)" in headers:
            set_cell(row, "Attendance Percentage (33.33%)",
                     f'=IFERROR(IF({col("Role")}{row}="Class-Naqeeb",'
                     f'{col("Present")}{row}/{col("No of Students")}{row}*0.33,0),0)')

        if "Direct Attendance Percentage" in headers:
            set_cell(row, "Direct Attendance Percentage",
                     f'=IFERROR(IF({col("Role")}{row}="Class-Naqeeb",'
                     f'{col("Present")}{row}/{col("No of Students")}{row},0),0)')

        if {"Ruju Percentage (33.33%)", "Fikr Percentage (33.33%)",
            "Attendance Percentage (33.33%)", "Overall  Percentage(100%)"} <= headers.keys():
            set_cell(row, "Overall  Percentage(100%)",
                     f"=IFERROR(SUM({col('Ruju Percentage (33.33%)')}{row}+"
                     f"{col('Fikr Percentage (33.33%)')}{row}+"
                     f"{col('Attendance Percentage (33.33%)')}{row}),0)")

        # Ensure all percentage formula columns have 0.00% number format
        for pct_col in ["Ruju Percentage (33.33%)", "Fikr Percentage (33.33%)",
                        "Attendance Percentage (33.33%)", "Direct Attendance Percentage",
                        "Overall  Percentage(100%)"]:
            if pct_col in headers:
                ws.cell(row=row, column=headers[pct_col]).number_format = "0.00%"

    # ---- extend Excel Table range (if Combined_Report is a Table) so structured
    #      references and filters/formatting cover the new rows.
    #      IMPORTANT: table.ref and table.autoFilter.ref must be updated together —
    #      a mismatch between them is what corrupts the file on open in Excel. ----
    new_last_row = next_row + len(ref_rows) - 1
    for tbl in ws.tables.values():
        min_col, min_row, max_col, max_row = range_boundaries(tbl.ref)
        if new_last_row > max_row:
            new_ref = f"{get_column_letter(min_col)}{min_row}:{get_column_letter(max_col)}{new_last_row}"
            tbl.ref = new_ref
            if tbl.autoFilter is not None:
                tbl.autoFilter.ref = new_ref

    wb.save(args.output)

    print(f"Appended {len(ref_rows)} row(s) ({next_row}-{new_last_row}) to '{args.output}'.\n")
    print("\n".join(log))


if __name__ == "__main__":
    main()