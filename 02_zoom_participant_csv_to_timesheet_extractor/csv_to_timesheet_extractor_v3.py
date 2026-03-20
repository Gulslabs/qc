import pandas as pd
import sys
from datetime import datetime, timedelta
import re
from openpyxl.styles import PatternFill, Font
import pdb

# Global Configuration
ENABLE_DURATION_THRESHOLD = False  # Set to True to enable duration threshold splitting
DURATION_THRESHOLD = 20  # Minimum duration to stay in main sheet
START_TIME_THRESHOLD_MINUTES = 15  # Minutes after start time to mark as late_joiner
END_TIME_THRESHOLD_MINUTES = 5    # Minutes before end time to mark as early_leaver
DATE_FORMAT= '%m-%d-%Y'
DATE_FORMAT_2 = '%d%b%Y'
H12_TIME_FORMAT = '%m/%d/%Y %I:%M:%S %p'

def process_name(name):
    """Process the name according to the specified rules"""
    # Handle NaN or None values
    if pd.isna(name) or name is None:
        return ""
    
    # Convert to string if it's not already
    name = str(name).strip()
    
    # Skip empty strings
    if not name:
        return ""
    
    # Rule 1: Simple name like 'Tariq Fareed' - copy as is
    if '(' not in name:
        processed_name = name
    else:
        # Rule 2: Check if first bracket contains 'Naqeeb'
        # Pattern: 'IK_Irfan Khan(Naqeeb) (Irfan Khan Yousufzai)'
        first_bracket_match = re.search(r'\(([^)]+)\)', name)
        if first_bracket_match:
            first_bracket_content = first_bracket_match.group(1)
            if 'Naqeeb' in first_bracket_content:
                # Keep everything up to and including the first bracket
                end_pos = first_bracket_match.end()
                processed_name = name[:end_pos]
            else:
                # Rule 3: If first bracket doesn't contain 'Naqeeb', remove everything from first bracket onwards
                start_pos = first_bracket_match.start()
                processed_name = name[:start_pos].strip()
        else:
            processed_name = name
    
    # Apply character replacement rules:
    # Replace hyphens with underscores
    processed_name = processed_name.replace('-', '_')    
    # Replace dots with underscores
    processed_name = processed_name.replace('.', '_')    
    # Replace one or more spaces with a single underscore
    processed_name = re.sub(r'\s+', '_', processed_name)
    
    # Consolidate multiple consecutive underscores into one
    processed_name = re.sub(r'_+', '_', processed_name)    
    return processed_name

def extract_date_from_datetime(datetime_str):
    """Extract date part from datetime string like '8/23/2025 07:35:40 PM'"""
    try:
        # Parse the datetime string and extract just the date part
        dt = datetime.strptime(datetime_str, H12_TIME_FORMAT)
        return dt.strftime(DATE_FORMAT)  # Return in same format
    except:
        # If parsing fails, try to extract date part manually
        if ' ' in datetime_str:
            return datetime_str.split(' ')[0]
        return datetime_str

def generate_filename(topic_cell, session_date):
    """Generate filename from the topic cell (A1)"""
    
    # Handle NaN or None values
    if pd.isna(topic_cell) or topic_cell is None:
        return "Time_Sheet_Unknown.xlsx"
    
    # Convert to string if it's not already
    topic_cell = str(topic_cell).strip()
    
    # Skip empty strings
    if not topic_cell:
        return "Time_Sheet_Unknown.xlsx"
    
    class_date = datetime.strptime(session_date, DATE_FORMAT).strftime(DATE_FORMAT_2).upper()    
    
    # Extract the part in parentheses for filename
    # '(KARMH-B02) #90 Surah Al-Balad' -> 'Time_Sheet_KARMH-B02_Surah Al-Balad.xlsx'
    
    # Remove parentheses and # symbol, replace spaces with underscores
    # First extract the content within parentheses
    match = re.search(r'\(([^)]+)\)', topic_cell)
    if match:
        code_part = match.group(1)  # KARMH-B02
    else:
        code_part = "Unknown"
    
    # Extract the title part (everything after the parentheses and #number)
    # Remove parentheses part and # with number
    remaining = re.sub(r'\([^)]+\)\s*#\d+\s*', '', topic_cell).strip()
    title_part = remaining.replace(' ', '_')    
    if class_date:
        filename = f"Time_Sheet_{code_part}_{title_part}_{class_date}.xlsx"   
    else:    
        filename = f"Time_Sheet_{code_part}_{title_part}.xlsx"
    print(f"Generated filename: {filename}")
    return filename

def split_data_by_duration_threshold(participants_data, duration_threshold=20):
    """
    Split participant data based on duration threshold
    
    Parameters:
    participants_data (list): List of participant dictionaries
    duration_threshold (int): Minimum duration to stay in main sheet
    
    Returns:
    tuple: (main_sheet_data, below_threshold_data)
    """
    main_sheet_data = []
    below_threshold_data = []
    
    for participant in participants_data:
        duration = participant.get('Duration (minutes)', '')
        
        # Check if duration is a number and below threshold
        try:
            duration_value = int(duration) if duration != '' else 0
            if duration_value < duration_threshold:
                below_threshold_data.append(participant)
            else:
                main_sheet_data.append(participant)
        except (ValueError, TypeError):
            # If duration is not a valid number, keep in main sheet
            main_sheet_data.append(participant)
    
    return main_sheet_data, below_threshold_data

def parse_halqa_ranges(halqa_ranges_str):
    """Parse halqa ranges string like '[001 - 030,571 - 585]' into list of tuples"""
    if not halqa_ranges_str or halqa_ranges_str.strip() == '':
        return []    
    ranges = []
    # Remove brackets and split by comma
    halqa_ranges_str = halqa_ranges_str.strip('[]')
    parts = halqa_ranges_str.split('|')
    
    for part in parts:
        part = part.strip()
        if '-' in part:
            start, end = part.split('-')
            start_num = int(start.strip())
            end_num = int(end.strip())
            ranges.append((start_num, end_num))    
    return ranges

def load_naqeeb_mapping(naqeeb_file_path='naqeeb_to_initial_TSAP-02.csv'):
    """
    Load Naqeeb mapping from CSV file
    
    Parameters:
    naqeeb_file_path (str): Path to the naqeeb mapping CSV file
    
    Returns:
    dict: Dictionary mapping initials to naqeeb names
    list: List of halqa mappings
    """
    naqeeb_mapping = {}  # initials -> (naqeeb_name, region)
    halqa_mapping = []   # list of (naqeeb_name, region, parsed_ranges)
    
    try:
        df = pd.read_csv(naqeeb_file_path)
        for _, row in df.iterrows():
            if len(row) >= 2:
                naqeeb_name = str(row.iloc[0]).strip()
                initials = str(row.iloc[1]).strip()
                region = str(row.iloc[2]).strip() if pd.notna(row.iloc[2]) else ''
                halqa_ranges_str = str(row.iloc[3]).strip() if pd.notna(row.iloc[3]) else ''                 
                naqeeb_mapping[initials] = (naqeeb_name, region)
                # Store halqa ranges if present
            if halqa_ranges_str and halqa_ranges_str.strip():
                ranges = parse_halqa_ranges(halqa_ranges_str)
                if ranges:
                    halqa_mapping.append((naqeeb_name, region, ranges))
        print(f"Loaded {len(naqeeb_mapping)} Naqeeb mappings from {naqeeb_file_path}")
    except FileNotFoundError:
        print(f"Warning: Naqeeb mapping file '{naqeeb_file_path}' not found. Naqeeb column will remain blank.")
    except Exception as e:
        print(f"Warning: Error loading Naqeeb mapping: {str(e)}")        
    return naqeeb_mapping, halqa_mapping

def extract_student_id(name):
    """Extract numeric ID from name (e.g., 'T1539', 'K313', '784')"""
    import re
    # Match patterns like T1539, K313, or standalone numbers
    match = re.search(r'[A-Za-z]?(\d+)', name)
    if match:
        return int(match.group(1))
    return None

def get_naqeeb_name_and_region(original_name, naqeeb_mapping, halqa_mapping=None):
    """
    Get Naqeeb name based on initials in the original name
    
    Parameters:
    original_name (str): The original name from CSV
    naqeeb_mapping (dict): Dictionary mapping initials to naqeeb names
    
    Returns:
    str: Naqeeb name if found, empty string otherwise
    """
    if not original_name or not naqeeb_mapping:
        return ''
    
    # If name contains 'Naqeeb', skip extraction
    if 'naqeeb' in original_name.lower():
        return '', ''
    
    # Convert original name to lowercase for comparison
    original_name_lower = original_name.lower()
    
    # Check if name starts with any initials pattern (with or without underscore)
    for initials, (naqeeb_name,region) in naqeeb_mapping.items():
        # Convert initials to lowercase for comparison
        initials_lower = initials.lower()
        # Check for "AZ_" pattern
        if original_name_lower.startswith(f"{initials_lower}_"):
            return naqeeb_name, region
        # Check for "RT " pattern (initials followed by space)
        elif initials_lower.startswith(f"{initials_lower} "):
            return naqeeb_name, region
        # # Check for exact initials at start followed by non-letter character
        elif len(original_name_lower) > len(initials_lower) and original_name_lower.startswith(initials_lower):
            # Make sure the character after initials is not a letter (to avoid partial matches)
            next_char = original_name_lower[len(initials_lower)]
            if not next_char.isalpha():
                return naqeeb_name, region
        # If no initials match found, try halqa range matching
        elif halqa_mapping:
            student_id = extract_student_id(original_name)
            if student_id:
                for naqeeb_name, region, halqa_ranges in halqa_mapping:
                    for start, end in halqa_ranges:
                        if start <= student_id <= end:
                            return naqeeb_name, region
    
    return '',''  # Not found

def consolidate_duplicate_participants(participants_data):
    """
    Consolidate participants with multiple join/leave sessions
    
    Parameters:
    participants_data (list): List of participant dictionaries
    
    Returns:
    list: Consolidated list with merged sessions for duplicate participants
    """   
    
    consolidated = {}
    
    for participant in participants_data:
        name = participant['Name']
        
        if name in consolidated:
            # Merge with existing entry
            existing = consolidated[name]
            
            # Parse times to find earliest join and latest leave
            try:
                # Current participant times
                #pdb.set_trace()
                #print(participant)
                current_join = datetime.strptime(participant['Join Time'], H12_TIME_FORMAT)
                current_leave = datetime.strptime(participant['Leave Time'], H12_TIME_FORMAT)
                
                # Existing participant times
                existing_join = datetime.strptime(existing['Join Time'], H12_TIME_FORMAT)
                existing_leave = datetime.strptime(existing['Leave Time'], H12_TIME_FORMAT)
                
                # Use earliest join time and latest leave time
                earliest_join = min(current_join, existing_join)
                latest_leave = max(current_leave, existing_leave)
                
                # Calculate total duration in minutes
                total_duration = int((latest_leave - earliest_join).total_seconds() / 60)
                
                # Add individual session duration to running total
                current_duration = int(participant['Duration (minutes)']) if participant['Duration (minutes)'] != '' else 0
                existing_session_duration = existing.get('session_duration', 0)
                total_session_duration = existing_session_duration + current_duration

                # Update consolidated entry
                consolidated[name].update({
                    'Join Time': earliest_join.strftime(H12_TIME_FORMAT),
                    'Leave Time': latest_leave.strftime(H12_TIME_FORMAT),
                    'Duration (minutes)': total_session_duration,  # Sum of all session durations
                })
                consolidated[name]['session_duration'] = total_session_duration
                
            except (ValueError, TypeError) as e:
                # If time parsing fails, keep the existing entry
                print(f"Warning: Could not parse times for {name}: {e}")
                continue
                
        else:
            # First occurrence of this participant
            duration = int(participant['Duration (minutes)']) if participant['Duration (minutes)'] != '' else 0
            participant_copy = participant.copy()
            participant_copy['session_duration'] = duration
            consolidated[name] = participant_copy
    
    # Convert back to list and remove the temporary session_duration field
    result = []
    for participant in consolidated.values():
        if 'session_duration' in participant:
            del participant['session_duration']
        result.append(participant)
    
    print(f"Consolidated {len(participants_data)} records into {len(result)} unique participants")
    return result

def convert_to_24hour_time(datetime_str):
    """
    Convert datetime string to 24-hour time format (HH:MM)
    
    Parameters:
    datetime_str (str): DateTime string like '8/23/2025 07:35:40 PM'
    
    Returns:
    str: Time in 24-hour format like '19:35' or original string if conversion fails
    """
    if not datetime_str or pd.isna(datetime_str):
        return ''
    
    try:
        # Parse the datetime string
        dt = datetime.strptime(str(datetime_str), H12_TIME_FORMAT)
        # Return only time in 24-hour format (HH:MM)
        return dt.strftime('%H:%M')
    except:
        # If parsing fails, return original string
        return str(datetime_str)

def convert_to_12hour_time(time_str):
    """
    Convert 24-hour time format to 12-hour time format with AM/PM
    
    Parameters:
    time_str (str): Time string in 24-hour format like '19:35' or '07:30'
    
    Returns:
    str: Time in 12-hour format like '07:35 PM' or '07:30 AM', or original string if conversion fails
    """
    if not time_str or pd.isna(time_str):
        return ''
    
    try:
        # Parse the time string (handles HH:MM format)
        dt = datetime.strptime(str(time_str), '%H:%M')
        # Return time in 12-hour format with AM/PM
        return dt.strftime('%I:%M %p')
    except:
        # If parsing fails, return original string
        return str(time_str)
    
def set_attendance_remarks(participants_data, session_start_time, session_end_time):
    """
    Set remarks based on join/leave time thresholds
    
    Parameters:
    participants_data (list): List of participant dictionaries
    session_start_time (str): Session start time in HH:MM format
    session_end_time (str): Session end time in HH:MM format
    
    Returns:
    list: Updated participants data with remarks
    """   
    
    try:
        # Parse session times
        start_time = datetime.strptime(session_start_time, '%H:%M')
        end_time = datetime.strptime(session_end_time, '%H:%M')
        
        # Calculate threshold times
        late_threshold = start_time + timedelta(minutes=START_TIME_THRESHOLD_MINUTES)
        early_threshold = end_time - timedelta(minutes=END_TIME_THRESHOLD_MINUTES)
        
        for participant in participants_data:
            remarks = []
                        
            try:
                # Parse participant times
                join_time = datetime.strptime(participant['Join Time'], '%H:%M')
                leave_time = datetime.strptime(participant['Leave Time'], '%H:%M')
                
                # Check if late joiner
                if join_time > late_threshold:
                    remarks.append('Late Joiner')
                
                # Check if early leaver
                if leave_time <= early_threshold:
                    remarks.append('Early Leaver')
                
                # Set remarks
                participant['Remarks'] = ', '.join(remarks)
                
            except (ValueError, TypeError):
                # If time parsing fails, skip this participant
                continue
                
    except (ValueError, TypeError):
        print(f"Warning: Could not parse session times '{session_start_time}' or '{session_end_time}'")
        return participants_data
    
    return participants_data

def apply_conditional_formatting(workbook, worksheet, participants_data, start_row=4):
    """
    Apply conditional formatting based on remarks
    
    Parameters:
    workbook: Excel workbook object
    worksheet: Excel worksheet object
    participants_data (list): List of participant dictionaries
    start_row (int): Starting row number for participant data
    """    
    
    # Define fill patterns and fonts
    amber_fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")  # Amber
    red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")    # Red
    black_font = Font(color="000000")  # Black text
    white_font = Font(color="FFFFFF")  # White text
    
    for idx, participant in enumerate(participants_data):
        row_num = start_row + idx
        remarks = participant.get('Remarks', '')
        
        if not remarks:
            continue
            
        # Check if both conditions are present
        if 'Late Joiner' in remarks and 'Early Leaver' in remarks:
            # Red background, white text for both Join Time (B) and Leave Time (C)
            join_cell = worksheet.cell(row=row_num, column=2)  # Column B - Join Time
            leave_cell = worksheet.cell(row=row_num, column=3)  # Column C - Leave Time
            
            join_cell.fill = red_fill
            join_cell.font = white_font
            leave_cell.fill = red_fill
            leave_cell.font = white_font
            
        else:
            # Individual conditions - amber background, black text
            if 'Late Joiner' in remarks:
                join_cell = worksheet.cell(row=row_num, column=2)  # Column B - Join Time
                join_cell.fill = amber_fill
                join_cell.font = black_font
                
            if 'Early Leaver' in remarks:
                leave_cell = worksheet.cell(row=row_num, column=3)  # Column C - Leave Time
                leave_cell.fill = amber_fill
                leave_cell.font = black_font

def convert_zoom_csv_to_timesheet(csv_file_path, start_time, end_time, naqeeb_mapping_file_path='naqeeb_to_initial_KARMH-02.csv'):
    """
    Convert Zoom CSV to Excel timesheet
    
    Parameters:
    csv_file_path (str): Path to the CSV file
    start_time (str): Start time for the session
    end_time (str): End time for the session
    """
    
    try:        
        # Read the CSV file
        meeting_info_df = pd.read_csv(csv_file_path, nrows=1)
        participants_df = pd.read_csv(csv_file_path, skiprows=2)
        
        # df = pd.read_csv(csv_file_path)
        # df = df.dropna(axis=1, how='all')  
        # Load Naqeeb mapping
        naqeeb_mapping, halqa_mapping  = load_naqeeb_mapping(naqeeb_mapping_file_path)
        
        # Get the topic from A1 (index 0, column 0)
        topic_cell = meeting_info_df.iloc[0, 0]  # A1 cell
        
        # Get the date from E1 (Start time column, row 1)
        start_time_cell = meeting_info_df.iloc[0, 4]  # E1 cell (Start time)
        
        # Handle NaN values for start_time_cell
        if pd.isna(start_time_cell):
            session_date = "Unknown"
        else:
            session_date = extract_date_from_datetime(str(start_time_cell))
        
         # Generate filename
        filename = generate_filename(topic_cell, session_date)
        # filename = "Test.xlsx"
        # Create the timesheet data
        # Based on the CSV structure:
        # Row 0: Topic and session info
        # Row 0: Headers ('Name (original name)', 'Email', etc.)
        # Row 1+: Participant data        
        data_start_idx = 1
        
        # Extract participant data
        participants_data = []
        
        for idx in range(data_start_idx, len(participants_df)):
            row = participants_df.iloc[idx]
            
            # Skip empty rows
            if pd.isna(row.iloc[0]) or str(row.iloc[0]).strip() == '':
                continue
            
            name = str(row.iloc[0]).strip()  # Name (original name)
            join_time = str(row.iloc[2]).strip() if pd.notna(row.iloc[2]) else ''  # Join time
            leave_time = str(row.iloc[3]).strip() if pd.notna(row.iloc[3]) else ''  # Leave time
            duration = str(row.iloc[4]).strip() if pd.notna(row.iloc[4]) else ''  # Duration (minutes)
            duration = int(duration) if duration.isdigit() else '' 
            
            # Process the name according to rules
            processed_name = process_name(name)
            
            # Get Naqeeb name based on original name
            naqeeb_name, region = get_naqeeb_name_and_region(name, naqeeb_mapping, halqa_mapping)
            
            
            participants_data.append({
                'Name': processed_name,
                'Join Time': join_time,
                'Leave Time': leave_time,
                'Duration (minutes)': duration,
                'Naqeeb Name': naqeeb_name,                
                'Remarks': '', # Leave blank
                'Region': region,  # Add region here
            })        
        # Consolidate duplicate participants
        participants_data = consolidate_duplicate_participants(participants_data)
        # Apply duration threshold splitting if enabled
        # Convert times to 24-hour format AFTER consolidation
        
        for participant in participants_data:
            participant['Join Time'] = convert_to_24hour_time(participant['Join Time'])
            participant['Leave Time'] = convert_to_24hour_time(participant['Leave Time'])
        
        
        participants_data = set_attendance_remarks(participants_data, start_time, end_time)        

        # Convert back to 12-hour format for output
        for participant in participants_data:
            participant['Join Time'] = convert_to_12hour_time(participant['Join Time'])
            participant['Leave Time'] = convert_to_12hour_time(participant['Leave Time'])

        if ENABLE_DURATION_THRESHOLD:
            main_sheet_data, below_threshold_data = split_data_by_duration_threshold(participants_data, DURATION_THRESHOLD)
            participants_df = pd.DataFrame(main_sheet_data)
            below_threshold_df = pd.DataFrame(below_threshold_data) if below_threshold_data else None
        else:
            participants_df = pd.DataFrame(participants_data)
            below_threshold_df = None
        
        # Create Excel file
        with pd.ExcelWriter(filename, engine='openpyxl') as writer:
            # Write participant data starting from row 3 (to leave space for header info)
            participants_df.to_excel(writer, sheet_name='Time_Sheet_Curated', index=False, header=True, startrow=2)
            
            # Get the workbook and worksheet to manually add header info
            workbook = writer.book
            worksheet = writer.sheets['Time_Sheet_Curated']
            
            # Write header information in row 1
            worksheet['A1'] = 'Date'
            worksheet['B1'] = session_date
            worksheet['C1'] = 'Start Time'
            worksheet['D1'] = start_time
            worksheet['E1'] = 'End Time'
            worksheet['F1'] = end_time
            # Apply conditional formatting
            apply_conditional_formatting(workbook, worksheet, participants_df.to_dict('records'))
            
            # Write below threshold data to Sheet2 if enabled and data exists
            if ENABLE_DURATION_THRESHOLD and below_threshold_df is not None and len(below_threshold_df) > 0:
                below_threshold_df.to_excel(writer, sheet_name='Below_20_Minutes', index=False, header=True, startrow=2)
                
                # Add header info to Sheet2 as well
                worksheet2 = writer.sheets['Below_20_Minutes']
                worksheet2['A1'] = 'Date'
                worksheet2['B1'] = session_date
                worksheet2['C1'] = 'Start Time'
                worksheet2['D1'] = start_time
                worksheet2['E1'] = 'End Time'
                worksheet2['F1'] = end_time
                apply_conditional_formatting(workbook, worksheet2, below_threshold_df.to_dict('records'))
        
        if ENABLE_DURATION_THRESHOLD:
            print(f"Successfully created timesheet: {filename}")
            print(f"Date: {session_date}")
            print(f"Start Time: {start_time}")
            print(f"End Time: {end_time}")
            print(f"Main sheet participants (>= {DURATION_THRESHOLD} minutes): {len(main_sheet_data)}")
            if below_threshold_data:
                print(f"Sheet2 participants (< {DURATION_THRESHOLD} minutes): {len(below_threshold_data)}")
        else:
            print(f"Successfully created timesheet: {filename}")
            print(f"Date: {session_date}")
            print(f"Start Time: {start_time}")
            print(f"End Time: {end_time}")
            print(f"Total participants processed: {len(participants_data)}")
        
        return filename
        
    except FileNotFoundError:
        print(f"Error: CSV file '{csv_file_path}' not found.")
        return None
    except Exception as e:
        print(f"Error processing file: {str(e)}")
        return None

def main():
    """Main function to handle command line arguments"""
    if len(sys.argv) != 5:
        print("Usage: python script.py <csv_file_path> <naqeeb_mapping_file_path> <start_time> <end_time>")
        print("Example: python script.py zoom_data.csv '07:30:00 PM' '11:00:00 PM'")
        sys.exit(1)
    
    csv_file_path = sys.argv[1]
    naqeeb_mapping_file_path = sys.argv[2]
    start_time = sys.argv[3]
    end_time = sys.argv[4]
    
    result = convert_zoom_csv_to_timesheet(csv_file_path, start_time, end_time, naqeeb_mapping_file_path)
    
    if result:
        print(f"Timesheet generated successfully: {result}")
    else:
        print("Failed to generate timesheet.")

if __name__ == "__main__":
    main()