import pandas as pd
import numpy as np
from datetime import datetime, timedelta
import sys
from openpyxl import load_workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.marker import DataPoint
from openpyxl.chart.shapes import GraphicalProperties
from openpyxl.chart.text import RichText
from openpyxl.drawing.text import CharacterProperties, ParagraphProperties, Paragraph
    
    

CUTOFF_TIME = datetime.strptime("23:20", "%H:%M")

def _style_chart_titles(chart):
    """Apply 16pt Times New Roman to chart title, x-axis title, y-axis title"""
    from openpyxl.chart.text import RichText
    from openpyxl.drawing.text import CharacterProperties, ParagraphProperties, Paragraph, Font

    def make_title_rich_text(text):
        cp = CharacterProperties(sz=1600, latin=Font(typeface="Times New Roman"))
        pp = ParagraphProperties(defRPr=cp)
        run = None
        from openpyxl.chart.text import Text
        from openpyxl.drawing.text import RegularTextRun
        run = RegularTextRun(rPr=cp, t=text)
        para = Paragraph(pPr=pp, r=[run])
        return RichText(p=[para])

    # Chart title
    if chart.title and chart.title.tx and chart.title.tx.rich:
        title_text = chart.title.tx.rich.p[0].r[0].t if chart.title.tx.rich.p[0].r else ""
        chart.title.tx.rich = make_title_rich_text(title_text)

    # X-axis title
    if chart.x_axis.title and chart.x_axis.title.tx and chart.x_axis.title.tx.rich:
        x_text = chart.x_axis.title.tx.rich.p[0].r[0].t if chart.x_axis.title.tx.rich.p[0].r else ""
        chart.x_axis.title.tx.rich = make_title_rich_text(x_text)

    # Y-axis title
    if chart.y_axis.title and chart.y_axis.title.tx and chart.y_axis.title.tx.rich:
        y_text = chart.y_axis.title.tx.rich.p[0].r[0].t if chart.y_axis.title.tx.rich.p[0].r else ""
        chart.y_axis.title.tx.rich = make_title_rich_text(y_text)

def _style_active_time_bars(chart, ws_data, last_row, count_col_letter):
    """Color the peak bar dark green, all others pale green"""
    from openpyxl.chart.marker import DataPoint
    from openpyxl.chart.shapes import GraphicalProperties

    counts = [ws_data[f'{count_col_letter}{row}'].value for row in range(4, last_row + 1)]
    max_count = max(counts)

    series = chart.series[0]
    for row in range(4, last_row + 1):
        value = ws_data[f'{count_col_letter}{row}'].value
        idx = row - 4
        color = "1E5B1E" if value == max_count else "C6E5C6"  # dark green / pale green
        pt = DataPoint(idx=idx)
        pt.graphicalProperties = GraphicalProperties(solidFill=color)
        series.data_points.append(pt)
    return series

def parse_time(time_str):
    """Convert time string (HH:MM or HH:MM AM/PM) to datetime object"""
    try:
        time_str = str(time_str).strip()
        # Try 24-hour format first
        try:
            return datetime.strptime(time_str, "%H:%M")
        except ValueError:
            # Fall back to 12-hour format with AM/PM
            return datetime.strptime(time_str, "%I:%M %p")
    except:
        print(f"Warning: Unable to parse time '{time_str}'. Expected HH:MM or HH:MM AM/PM.")
        return None
    
def time_to_decimal(dt):
    """Convert datetime to decimal (Excel format)"""
    if pd.isna(dt) or dt is None:
        return None
    return (dt.hour * 3600 + dt.minute * 60 + dt.second) / 86400

def decimal_to_time_str(decimal):
    """Convert decimal back to time string"""
    if pd.isna(decimal):
        return ""
    total_seconds = decimal * 86400
    hours = int(total_seconds // 3600)
    minutes = int((total_seconds % 3600) // 60)
    return f"{hours:02d}:{minutes:02d}"

def create_time_bins(min_time, max_time, interval_minutes):
    """Create time bins from min to max with given interval"""
    bins_start = []
    bins_end = []
    
    current = min_time
    
    while current <= max_time:
        next_time = current + timedelta(minutes=interval_minutes)
        bins_start.append(current)
        bins_end.append(next_time)
        current = next_time
        
        # Stop at first occurrence after CUTOFF_TIME
        if next_time > CUTOFF_TIME:
            break
    
    return bins_start, bins_end

def calculate_frequency(times, bin_ends):
    """Calculate frequency distribution like Excel FREQUENCY function"""
    counts = []
    time_decimals = [time_to_decimal(t) for t in times if t is not None]
    bin_end_decimals = [time_to_decimal(b) for b in bin_ends]
    
    for i, bin_end in enumerate(bin_end_decimals):
        if i == 0:
            # First bin: count values <= bin_end
            count = sum(1 for t in time_decimals if t <= bin_end)
        else:
            # Other bins: count values > previous bin_end and <= current bin_end
            count = sum(1 for t in time_decimals if bin_end_decimals[i-1] < t <= bin_end)
        counts.append(count)
    
    return counts

def process_excel_file(file_path, interval_minutes=5):
    """
    Process Excel file and add time bin analysis columns
    
    Parameters:
    - file_path: Path to the Excel file
    - interval_minutes: Time interval for bins (default: 5 minutes)
    """
    
    # Read the Excel file
    # Row 1-2: Metadata, Row 3: Headers, Row 4+: Data
    print(f"Reading Excel file: {file_path}")
    df = pd.read_excel(file_path, header=2)  # Row 3 (index 2) is the header
    
    # Find the last row with data
    last_row = df.dropna(subset=['Name']).index[-1] + 1  # +1 for 0-indexing
    print(f"Data found up to row: {last_row + 1}")  # +1 for Excel row number
    
    # Parse Join Time and Leave Time
    print("Parsing time columns...")
    join_times = [parse_time(t) for t in df['Join Time'].dropna()]
    leave_times = [parse_time(t) for t in df['Leave Time'].dropna()]
    # print(f"Parsed {join_times} Join Times and {leave_times} Leave Times.")
    # Filter out None values
    join_times = [t for t in join_times if t is not None]
    leave_times = [t for t in leave_times if t is not None]
    # print(f"Parsed {len(join_times)} Join Times and {len(leave_times)} Leave Times.")
    if not join_times or not leave_times:
        print("Error: No valid time data found!")
        return
    
    # Find min and max times
    min_join = min(join_times)
    max_join = max(join_times)
    min_leave = min(leave_times)
    max_leave = max(leave_times)
    
    print(f"Join time range: {min_join.strftime('%H:%M')} to {max_join.strftime('%H:%M')}")
    print(f"Leave time range: {min_leave.strftime('%H:%M')} to {max_leave.strftime('%H:%M')}")
    
    # Create time bins for Join Time
    print(f"\nCreating {interval_minutes}-minute bins for Join Time...")
    join_bins_start, join_bins_end = create_time_bins(min_join, max_join, interval_minutes)
    
    # Create time bins for Leave Time
    print(f"Creating {interval_minutes}-minute bins for Leave Time...")
    leave_bins_start, leave_bins_end = create_time_bins(min_leave, max_leave, interval_minutes)
    
    # Calculate frequencies
    print("Calculating student counts...")
    join_counts = calculate_frequency(join_times, join_bins_end)
    leave_counts = calculate_frequency(leave_times, leave_bins_end)
    
    # Create new columns in DataFrame
    max_bins = max(len(join_bins_start), len(leave_bins_start))
    
    # Initialize columns with empty values
    df['Bin_Start_Join_Time'] = ''
    df['Bin_End_Join_Time'] = ''
    df['Join_Time_Range'] = ''
    df['Join_Student_Count'] = ''
    df['Bin_Start_Leave_Time'] = ''
    df['Bin_End_Leave_Time'] = ''
    df['Leave_Time_Range'] = ''
    df['Leave_Student_Count'] = ''
    
    # Convert count columns to Int64 (nullable integer) to allow integer assignment
    df['Join_Student_Count'] = pd.to_numeric(df['Join_Student_Count'], errors='coerce').astype('Int64')
    df['Leave_Student_Count'] = pd.to_numeric(df['Leave_Student_Count'], errors='coerce').astype('Int64')
    
    # Fill in the bin data starting from row 0 in DataFrame (which is row 4 in Excel)
    # Since we read with header=2, the DataFrame starts fresh from 0
    start_row = 0  # First data row in DataFrame (Row 4 in Excel)
    
    for i in range(len(join_bins_start)):
        if start_row + i < len(df):
            start_str = join_bins_start[i].strftime('%H:%M')
            end_str = join_bins_end[i].strftime('%H:%M')
            df.at[start_row + i, 'Bin_Start_Join_Time'] = start_str
            df.at[start_row + i, 'Bin_End_Join_Time'] = end_str
            df.at[start_row + i, 'Join_Time_Range'] = f"{start_str} - {end_str}"
            df.at[start_row + i, 'Join_Student_Count'] = join_counts[i]
    
    for i in range(len(leave_bins_start)):
        if start_row + i < len(df):
            start_str = leave_bins_start[i].strftime('%H:%M')
            end_str = leave_bins_end[i].strftime('%H:%M')
            df.at[start_row + i, 'Bin_Start_Leave_Time'] = start_str
            df.at[start_row + i, 'Bin_End_Leave_Time'] = end_str
            df.at[start_row + i, 'Leave_Time_Range'] = f"{start_str} - {end_str}"
            df.at[start_row + i, 'Leave_Student_Count'] = leave_counts[i]
    
    # Save to new Excel file
    output_file = file_path.replace('.xlsx', '_with_report.xlsx')
    print(f"\nSaving results to: {output_file}")
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Time_Sheet_Curated', index=False, header=True, startrow=2)
    
    # Print summary
    print("\n" + "="*60)
    print("SUMMARY")
    print("="*60)
    print(f"Join Time Bins (Total: {len(join_bins_start)})")
    # print(f"{'Time Range':<20} {'Student Count':<15}")
    # print("-" * 35)
    for i in range(len(join_bins_start)):
        time_range = f"{join_bins_start[i].strftime('%H:%M')}-{join_bins_end[i].strftime('%H:%M')}"
        # print(f"{time_range:<20} {join_counts[i]:<15}")
    
    print(f"Leave Time Bins (Total: {len(leave_bins_start)})")
    # print(f"{'Time Range':<20} {'Student Count':<15}")
    # print("-" * 35)
    for i in range(len(leave_bins_start)):
        time_range = f"{leave_bins_start[i].strftime('%H:%M')}-{leave_bins_end[i].strftime('%H:%M')}"
        # print(f"{time_range:<20} {leave_counts[i]:<15}")
    
    print(f"✓ Processing complete! Output saved to: {output_file}")
    return df

def add_active_student_time_series(file_path, interval_minutes=3):
    """Add Active_Time and Active_Student_Count columns (P, Q) to the report"""
    from openpyxl import load_workbook
    
    wb = load_workbook(file_path)
    ws_data = wb['Time_Sheet_Curated']
    
    # Locate Name / Join Time / Leave Time columns by header (row 3)
    headers = {ws_data.cell(row=3, column=c).value: c for c in range(1, ws_data.max_column + 1)}
    name_col = headers.get('Name')
    join_col = headers.get('Join Time')
    leave_col = headers.get('Leave Time')
    
    if not (name_col and join_col and leave_col):
        print("Error: Could not locate Name/Join Time/Leave Time columns.")
        return
    
    # Collect (join, leave) pairs per student
    students = []
    for row in range(4, ws_data.max_row + 1):
        if ws_data.cell(row=row, column=name_col).value is None:
            continue
        jt = parse_time(ws_data.cell(row=row, column=join_col).value)
        lt_val = ws_data.cell(row=row, column=leave_col).value
        lt = parse_time(lt_val) if lt_val is not None else None
        if jt is not None:
            students.append((jt, lt))
    
    if not students:
        print("No valid Join/Leave data found for active time series.")
        return
    
    # Generate active counts from min join time to CUTOFF_TIME
    min_join = min(jt for jt, lt in students)
    active_times, active_counts = [], []
    current = min_join
    while current <= CUTOFF_TIME:
        count = sum(1 for jt, lt in students if jt <= current and (lt is None or lt > current))
        active_times.append(current)
        active_counts.append(count)
        current += timedelta(minutes=interval_minutes)
    
    # Write to columns P (16) and Q (17)
    ws_data['P3'] = 'Active_Time'
    ws_data['Q3'] = 'Active_Student_Count'
    for i, (t, c) in enumerate(zip(active_times, active_counts)):
        row = 4 + i
        ws_data[f'P{row}'] = t.strftime('%H:%M')
        ws_data[f'Q{row}'] = c

    wb.save(file_path)
    print(f"✓ Active time series added ({len(active_times)} points)")
    

def add_join_time_chart(file_path):
    """Add 'Early & Late Joiners' chart to a new worksheet"""
    
    wb = load_workbook(file_path)
    ws_data = wb['Time_Sheet_Curated']
    
    join_last_row = 0
    values = []
    for row in range(4, ws_data.max_row + 1):
        val = ws_data[f'K{row}'].value  # Join_Student_Count
        if val is not None:
            join_last_row = row
            values.append(val)
    
    if join_last_row > 3:
        ws_join = wb.create_sheet('Early & Late Joiners')
        
        chart = BarChart()
        chart.type = "col"
        chart.title = "Early & Late Join"
        chart.y_axis.title = "Student Count"
        chart.x_axis.title = "Time"     
        _style_chart_titles(chart)   
        chart.legend = None
        chart.height = 15
        chart.width = 35
        chart.gapWidth = 18
        
        labels = Reference(ws_data, min_col=10, min_row=4, max_row=join_last_row)  # J: Join_Time_Range
        data = Reference(ws_data, min_col=11, min_row=3, max_row=join_last_row)    # K: Join_Student_Count
        
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(labels)
        
        series = chart.series[0]
        series.dLbls = DataLabelList()
        series.dLbls.showVal = True
        series.dLbls.showSerName = False
        series.dLbls.showCatName = False
        series.dLbls.showLegendKey = False
        series.dLbls.showPercent = False
        series.dLbls.showBubbleSize = False
        series.dLbls.numFmt = "0"
        
        cp = CharacterProperties(sz=800, b=False)
        pp = ParagraphProperties(defRPr=cp)
        series.dLbls.txPr = RichText(p=[Paragraph(pPr=pp, endParaRPr=cp)])
        
        sorted_vals = sorted(values, reverse=True)
        top_n = max(1, int(len(sorted_vals) * 0.15))
        threshold = sorted_vals[top_n - 1]
        
        for row in range(4, join_last_row + 1):
            value = ws_data[f'K{row}'].value
            idx = row - 4
            color = "70AD47" if (value and value >= threshold) else "C00000"
            pt = DataPoint(idx=idx)
            pt.graphicalProperties = GraphicalProperties(solidFill=color)
            series.data_points.append(pt)
        
        chart.x_axis.textRotation = -90
        chart.x_axis.delete = False
        x_cp = CharacterProperties(sz=900)
        x_pp = ParagraphProperties(defRPr=x_cp)
        chart.x_axis.txPr = RichText(p=[Paragraph(pPr=x_pp, endParaRPr=x_cp)])
        
        ws_join.add_chart(chart, "A1")
        wb.save(file_path)
        print(f"✓ Join time chart added")


def add_leave_time_chart(file_path):
    """Add 'Early Left Graph' chart to a new worksheet"""   
    
    wb = load_workbook(file_path)
    ws_data = wb['Time_Sheet_Curated']
    
    leave_last_row = 0
    values = []
    for row in range(4, ws_data.max_row + 1):
        val = ws_data[f'O{row}'].value  # Leave_Student_Count
        if val is not None:
            leave_last_row = row
            values.append(val)
    
    if leave_last_row > 3:
        ws_leave = wb.create_sheet('Early Left Graph')
        
        chart = BarChart()
        chart.type = "col"
        chart.title = "Early Left Graph"
        chart.y_axis.title = "Student Count"
        chart.x_axis.title = "Time"
        _style_chart_titles(chart)
        chart.legend = None
        chart.height = 15
        chart.width = 35
        chart.gapWidth = 18
        
        labels = Reference(ws_data, min_col=14, min_row=4, max_row=leave_last_row)  # N: Leave_Time_Range
        data = Reference(ws_data, min_col=15, min_row=3, max_row=leave_last_row)    # O: Leave_Student_Count
        
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(labels)
        
        series = chart.series[0]
        series.dLbls = DataLabelList()
        series.dLbls.showVal = True
        series.dLbls.showSerName = False
        series.dLbls.showCatName = False
        series.dLbls.showLegendKey = False
        series.dLbls.showPercent = False
        series.dLbls.showBubbleSize = False
        series.dLbls.numFmt = "0"
        
        cp = CharacterProperties(sz=700, b=False)
        pp = ParagraphProperties(defRPr=cp)
        series.dLbls.txPr = RichText(p=[Paragraph(pPr=pp, endParaRPr=cp)])
        
        sorted_vals = sorted(values, reverse=True)
        top_n = max(1, int(len(sorted_vals) * 0.15))
        threshold = sorted_vals[top_n - 1]
        
        for row in range(4, leave_last_row + 1):
            value = ws_data[f'O{row}'].value
            idx = row - 4
            color = "70AD47" if (value and value >= threshold) else "C00000"
            pt = DataPoint(idx=idx)
            pt.graphicalProperties = GraphicalProperties(solidFill=color)
            series.data_points.append(pt)
        
        chart.x_axis.textRotation = -90
        chart.x_axis.delete = False
        x_cp = CharacterProperties(sz=900)
        x_pp = ParagraphProperties(defRPr=x_cp)
        chart.x_axis.txPr = RichText(p=[Paragraph(pPr=x_pp, endParaRPr=x_cp)])
        
        ws_leave.add_chart(chart, "A1")
        wb.save(file_path)
        print(f"✓ Leave time chart added")

def add_active_time_chart(file_path):
    """Add 'Active Students Over Time' chart to a new worksheet"""
    from openpyxl import load_workbook
    from openpyxl.chart import BarChart, Reference
    from openpyxl.chart.label import DataLabelList
    from openpyxl.chart.text import RichText
    from openpyxl.drawing.text import CharacterProperties, ParagraphProperties, Paragraph
    
    wb = load_workbook(file_path)
    ws_data = wb['Time_Sheet_Curated']
    
    active_last_row = 0
    for row in range(4, ws_data.max_row + 1):
        if ws_data[f'Q{row}'].value is not None:
            active_last_row = row
    
    if active_last_row > 3:
        ws_active = wb.create_sheet('Active Students Over Time')
        
        chart = BarChart()
        chart.type = "col"
        chart.title = "Active Students Over Time"
        chart.y_axis.title = "Student Count"
        chart.x_axis.title = "Time"
        _style_chart_titles(chart)
        chart.legend = None
        chart.height = 10
        chart.width = 28
        chart.gapWidth = 20
        
        labels = Reference(ws_data, min_col=16, min_row=4, max_row=active_last_row)  # P: Active_Time
        data = Reference(ws_data, min_col=17, min_row=3, max_row=active_last_row)    # Q: Active_Student_Count
        
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(labels)
                
        series = _style_active_time_bars(chart, ws_data, active_last_row, 'Q')
        series.dLbls = DataLabelList()
        series.dLbls.showVal = True
        series.dLbls.showSerName = False
        series.dLbls.showCatName = False
        series.dLbls.showLegendKey = False
        series.dLbls.showPercent = False
        series.dLbls.showBubbleSize = False
        series.dLbls.numFmt = "0"
        
        cp = CharacterProperties(sz=700, b=False)
        pp = ParagraphProperties(defRPr=cp)
        series.dLbls.txPr = RichText(p=[Paragraph(pPr=pp, endParaRPr=cp)])
        
        chart.x_axis.textRotation = -90   # vertical
        chart.x_axis.delete = False
        x_cp = CharacterProperties(sz=900)
        x_pp = ParagraphProperties(defRPr=x_cp)
        chart.x_axis.txPr = RichText(p=[Paragraph(pPr=x_pp, endParaRPr=x_cp)])
        
        ws_active.add_chart(chart, "A1")
        wb.save(file_path)
        print(f"✓ Active time chart added")

# Main execution
if __name__ == "__main__":
    # Example usage
    if len(sys.argv) < 2:
        print("Usage: python script.py <excel_file_path> [interval_minutes]")
        print("Example: python script.py attendance.xlsx 5")
        print("\nOr modify the script and run directly:")
        print("  file_path = 'your_file.xlsx'")
        print("  interval_minutes = 5")
        sys.exit(1)
    
    file_path = sys.argv[1]
    interval_minutes = int(sys.argv[2]) if len(sys.argv) > 2 else 5    
    process_excel_file(file_path, interval_minutes)
    output_file = file_path.replace('.xlsx', '_with_report.xlsx')
    add_join_time_chart(output_file)
    add_leave_time_chart(output_file)
    add_active_student_time_series(output_file, interval_minutes)
    add_active_time_chart(output_file)