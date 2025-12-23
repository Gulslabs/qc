import os
import shutil
from pathlib import Path
from datetime import datetime
from openpyxl import load_workbook, Workbook
from typing import Dict, List, Tuple, Optional
import warnings

# Suppress openpyxl warnings about unsupported Excel features
warnings.filterwarnings('ignore', category=UserWarning, module='openpyxl')


class AttendanceMasterGenerator:
    """
    Generates a master attendance sheet from individual Naqeeb sheets.
    Filters records by ID prefix and creates a consolidated output file.
    """
    
    def __init__(self, parent_folder: str, attendance_date: str, id_prefix: str, target_sheet_name: str = "KARMH-B02"):
        """
        Initialize the attendance generator.
        
        Args:
            parent_folder: Path to folder containing source files
            attendance_date: Date column to extract (e.g., '20-12-25')
            id_prefix: ID prefix to filter (e.g., 'KARMH-B02-G1401')
        """
        self.parent_folder = Path(parent_folder)
        self.attendance_date = attendance_date
        self.id_prefix = id_prefix
        self.output_file = self.parent_folder / f"master_sheet_{attendance_date}.xlsx"
        self.processed_folder = self.parent_folder / "Processed"
        self.target_sheet_name = target_sheet_name
        # Create processed folder if it doesn't exist
        self.processed_folder.mkdir(exist_ok=True)
        
        # Valid attendance values
        self.valid_values = {'-', 'P', 'A', 'L', 'D'}
        
        # Track skipped records
        self.skipped_records = []
        
    def get_source_files(self) -> List[Path]:
        """Get all Excel files in the parent folder."""
        files = []
        for file in self.parent_folder.glob("*.xlsx"):
            if not file.name.startswith("~$") and not file.name.startswith("master_sheet"):
                files.append(file)
        return files
    
    def find_column_index(self, ws, header_row: int, column_name: str) -> Optional[int]:
        """Find the column index for a given column name (case-insensitive)."""
        for col_idx, cell in enumerate(ws[header_row], start=1):
            if cell.value:
                cell_value = str(cell.value).strip()
                if cell_value.lower() == column_name.lower():
                    return col_idx
        return None
    
    def find_date_column(self, ws, header_row: int, date_str: str) -> Optional[int]:
        """
        Find date column by matching text strings and Excel date objects.
        Handles both short (DD-MM-YY) and long (DD-MM-YYYY) formats.
        """
        from datetime import datetime as dt
        
        # Parse the input date string
        parts = date_str.split('-')
        if len(parts) != 3:
            return None
        
        day, month, year = parts
        
        # Create both short and long format variations
        if len(year) == 2:
            year_full = f"20{year}"
            year_short = year
        else:
            year_full = year
            year_short = year[-2:]
        
        short_date = f"{day}-{month}-{year_short}"
        long_date = f"{day}-{month}-{year_full}"
        
        # Try to parse as datetime for comparison with Excel date objects
        try:
            target_date = dt(int(year_full), int(month), int(day))
        except:
            target_date = None
        
        # Search through all columns
        for col_idx in range(1, ws.max_column + 1):
            cell = ws.cell(header_row, col_idx)
            if cell.value:
                # Check if it's a string match
                if isinstance(cell.value, str):
                    cell_value = str(cell.value).strip()
                    if cell_value in [short_date, long_date]:
                        return col_idx
                
                # Check if it's an Excel date object
                elif isinstance(cell.value, dt):
                    if target_date and cell.value.date() == target_date.date():
                        return col_idx
        
        return None
    
    def extract_records_from_source(self, source_file: Path) -> List[Dict]:
        """
        Extract attendance records from a source file.
        Finds the worksheet named 'KARMH-B02' with header at row 3.
        Only includes records where ID starts with the specified prefix.
        
        Returns:
            List of records with ID, Region, Naqeeb, and Attendance
        """
        records = []
        header_row = 3  # Fixed header row for source files        
        
        try:
            wb = load_workbook(source_file, read_only=True, data_only=True)
            
            # Find the worksheet named 'KARMH-B02'
            if self.target_sheet_name not in wb.sheetnames:
                print(f"  Warning: Worksheet '{self.target_sheet_name}' not found in {source_file.name}")
                print(f"  Available sheets: {', '.join(wb.sheetnames)}")
                wb.close()
                return records
            
            ws = wb[self.target_sheet_name]
            # print(f"  Processing worksheet: '{target_sheet_name}'")            
            # DEBUG: Print headers from row 3
            # print(f"  Headers at row {header_row}:")
            # for col_idx in range(1, min(15, ws.max_column + 1)):
            #     cell = ws.cell(header_row, col_idx)
            #     if cell.value:
            #         print(f"    Col {col_idx}: '{cell.value}'")
            
            # Find column indices
            id_col = self.find_column_index(ws, header_row, "Assigned ID#")
            region_col = self.find_column_index(ws, header_row, "Region")
            naqeeb_col = self.find_column_index(ws, header_row, "Naqeeb")
            attendance_col = self.find_date_column(ws, header_row, self.attendance_date)
            
            print(f"  Column search: ID#={id_col}, Region={region_col}, Naqeeb={naqeeb_col}, {self.attendance_date}={attendance_col}")
            
            if not all([id_col, region_col, naqeeb_col]):
                print(f"  Warning: Required columns not found in {source_file.name}")
                wb.close()
                return records
            
            if not attendance_col:
                print(f"  Warning: Attendance date column '{self.attendance_date}' not found in {source_file.name}")
                wb.close()
                return records
            
            # Extract data from rows (starting after header row)
            for row_idx in range(header_row + 1, ws.max_row + 1):
                id_val = ws.cell(row_idx, id_col).value
                region_val = ws.cell(row_idx, region_col).value
                naqeeb_val = ws.cell(row_idx, naqeeb_col).value
                attendance_val = ws.cell(row_idx, attendance_col).value
                
                # Skip if key values are missing
                if not all([id_val, region_val, naqeeb_val]):
                    continue
                
                # Clean and normalize values
                id_val = str(id_val).strip()
                region_val = str(region_val).strip()
                naqeeb_val = str(naqeeb_val).strip()
                
                # Check if ID starts with the specified prefix
                if id_val.startswith(self.id_prefix):
                    # Clean attendance value
                    attendance_str = '-'
                    if attendance_val:
                        attendance_str = str(attendance_val).strip().upper()
                        if attendance_str not in self.valid_values:
                            attendance_str = '-'
                    
                    records.append({
                        'id': id_val,
                        'region': region_val,
                        'naqeeb': naqeeb_val,
                        'attendance': attendance_str,
                        'source_file': source_file.name,
                        'row': row_idx
                    })
                else:
                    # Track skipped records
                    self.skipped_records.append({
                        'id': id_val,
                        'region': region_val,
                        'naqeeb': naqeeb_val,
                        'source_file': source_file.name,
                        'row': row_idx,
                        'reason': f"ID doesn't start with '{self.id_prefix}'"
                    })
            
            wb.close()
            print(f"  Extracted {len(records)} matching records")
            
        except Exception as e:
            print(f"  Error processing {source_file.name}: {str(e)}")
        
        return records
    
    def create_master_sheet(self, all_records: List[Dict]):
        """Create the master attendance sheet with collected records."""
        try:
            # Create new workbook in write_only mode
            wb = Workbook(write_only=True)
            ws = wb.create_sheet("Master Attendance")
            
            # Write header row
            headers = ["ID# Assigned", "Region", "Naqeebs", self.attendance_date]
            ws.append(headers)
            
            # Write data rows in batches for better performance
            for record in all_records:
                ws.append([
                    record['id'],
                    record['region'],
                    record['naqeeb'],
                    record['attendance']
                ])
            
            # Save the file
            wb.save(self.output_file)
            wb.close()
            
            print(f"\n✓ Master sheet created: {self.output_file.name}")
            print(f"  Total records written: {len(all_records)}")
            
        except Exception as e:
            print(f"\n✗ Error creating master sheet: {str(e)}")
            import traceback
            traceback.print_exc()
    
    def move_to_processed(self, source_file: Path):
        """Move processed file to Processed folder with timestamp."""
        try:
            timestamp = datetime.now().strftime("%H_%M")
            file_stem = source_file.stem
            file_ext = source_file.suffix
            new_name = f"{file_stem}_{timestamp}{file_ext}"
            destination = self.processed_folder / new_name
            
            shutil.move(str(source_file), str(destination))
            print(f"  Moved to Processed/{new_name}")
            
        except Exception as e:
            print(f"  Error moving {source_file.name}: {str(e)}")
    
    def print_skipped_records(self):
        """Print all skipped records for review."""
        if not self.skipped_records:
            print("\n✓ No records were skipped (all IDs matched the prefix)")
            return
        
        print(f"\n{'='*80}")
        print(f"SKIPPED RECORDS: {len(self.skipped_records)} record(s)")
        print(f"{'='*80}")
        
        # Group by source file
        by_file = {}
        for record in self.skipped_records:
            file_name = record['source_file']
            if file_name not in by_file:
                by_file[file_name] = []
            by_file[file_name].append(record)
        
        for file_name, records in by_file.items():
            print(f"\nFile: {file_name}")
            print("-" * 80)
            for record in records[:20]:  # Show first 20 per file
                print(f"  Row {record['row']}: ID={record['id']}, Region={record['region']}, Naqeeb={record['naqeeb']}")
            if len(records) > 20:
                print(f"  ... and {len(records) - 20} more records from this file")
        
        print(f"\n{'='*80}")
    
    def process_all(self):
        """Main processing method to generate master attendance sheet."""
        start_time = datetime.now()
        
        print(f"\n{'='*60}")
        print(f"Attendance Master Sheet Generator")
        print(f"Started at: {start_time.strftime('%Y-%m-%d %H:%M:%S')}")
        print(f"{'='*60}")
        print(f"Parent Folder:    {self.parent_folder}")
        print(f"Attendance Date:  {self.attendance_date}")
        print(f"ID Prefix Filter: {self.id_prefix}")
        print(f"Output File:      {self.output_file.name}")
        print(f"{'='*60}\n")
        
        # Get all source files
        source_files = self.get_source_files()
        
        if not source_files:
            print("No source files found to process.")
            return
        
        print(f"Found {len(source_files)} source file(s) to process\n")
        
        # Collect all records
        all_records = []
        
        for source_file in source_files:
            print(f"Processing: {source_file.name}")
            
            records = self.extract_records_from_source(source_file)
            all_records.extend(records)
            
            # Move to processed folder
            #self.move_to_processed(source_file)
            print()
        
        # Create master sheet
        print(f"{'='*60}")
        print("Creating Master Sheet")
        print(f"{'='*60}")
        
        if all_records:
            self.create_master_sheet(all_records)
        else:
            print("\n⚠️  No matching records found. Master sheet not created.")
        
        # Print skipped records
        self.print_skipped_records()
        
        # Print summary
        end_time = datetime.now()
        duration = end_time - start_time
        duration_seconds = duration.total_seconds()
        duration_minutes = duration_seconds / 60
        
        print(f"\n{'='*60}")
        print(f"Process Complete!")
        print(f"{'='*60}")
        print(f"Started at:  {start_time.strftime('%Y-%m-%d %H:%M:%S')}")
        print(f"Ended at:    {end_time.strftime('%Y-%m-%d %H:%M:%S')}")
        print(f"Duration:    {duration_minutes:.2f} minutes ({duration_seconds:.2f} seconds)")
        print(f"\nRecords included: {len(all_records)}")
        print(f"Records skipped:  {len(self.skipped_records)}")
        print(f"{'='*60}\n")


def main():
    """
    Usage: python attendance_generator.py <parent_folder> <attendance_date> <id_prefix>
    """
    import sys
    
    # Check command line arguments
    if len(sys.argv) != 4:
        print("Error: Invalid number of arguments")
        print("Usage: python attendance_generator.py <parent_folder> <attendance_date> <id_prefix>")
        print("Example: python attendance_generator.py /path/to/folder 20-12-25 KARMH-B02")
        sys.exit(1)
    
    # Get arguments
    parent_folder = sys.argv[1]
    attendance_date = sys.argv[2]
    id_prefix = sys.argv[3]
    
    # Validate parent folder exists
    if not os.path.exists(parent_folder):
        print(f"Error: Parent folder does not exist: {parent_folder}")
        sys.exit(1)
    
    # Create generator and process
    generator = AttendanceMasterGenerator(parent_folder, attendance_date, id_prefix, id_prefix)
    generator.process_all()


if __name__ == "__main__":
    main()